import os,json, collections
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


# Prepare styles
font_white = Font(color="FFFFFF")
pattern_green = PatternFill("solid", start_color="70AD47")
pattern_orange = PatternFill("solid", start_color="ED7D31")

# Initialisation of the xlsx file
filename = "AIrating_evaluation_2.xlsx"

# Create the workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "Instructions"

# Set the width of the column
sheet.column_dimensions['A'].width = 100

sheet['A1'].font = Font(bold=True)
sheet["A1"] = "Evaluation AI rating - Impaakt - HESSO"
sheet["A2"] = "Version : 2.0"
sheet["A3"] = "Date : 2024-01-26"
sheet['A4'].alignment = Alignment(wrap_text=True)
sheet["A4"] = "Instructions:\n- Original queries are highlighted in yellow at the top of the page.\n- In each of the following sheets, you will find two tables of different colours.\n- Each table provides the result of a strategy for a given query, but positions A or B have been randomly assigned between the sheets.\n- For each table, please test all the URLs in descending order and fill in the cells on the right:\n    - in the first column indicate with \"YES\" or \"NO\" whether the page is accessible;\n    - in the second column indicate with \"YES\" or \"NO\" whether the page is relevant.\n- For each table, you can stop after the first relevant result."


# Initialisation of data validation for response cells
dv = DataValidation(type="list", formula1='"YES,NO"')
# Optionally set a custom error message
#dv.error ='Your entry is not in the list'
#dv.errorTitle = 'Invalid Entry'


# Process the data
rfile = open("input/20231228_company_topic_queries.csv",encoding="utf-8")
i = 1
for line in rfile :
    i_str = str(i)
    query = line.strip()
    sources = collections.OrderedDict()
    sources_reranked = collections.OrderedDict()

    #get data from Google API
    with open("output/2024-01/1_Google/"+i_str+".json",encoding="utf-8") as file:
       data = json.loads(file.read())
       for org in data["organic_results"] :
           url = org["link"]
           sources[url] = {}
           sources[url]["organic_position"] = org["position"]

    # Get data from Impaakt webservice processing
    with open("output/2024-01/4_impaakt/"+i_str+".json",encoding="utf-8") as file:
       data = json.loads(file.read())
       for org2 in data["source"] :
           url = org2["url"]
           sources_reranked[url] = {}
           if ("impaakt" in org2) :
               sources_reranked[url]["impaakt"] = org2["impaakt"]
           else :
               sources_reranked[url]["impaakt"] = "0"
    sorted_sources = dict(sorted(sources_reranked.items(), key=lambda item: float(item[1]["impaakt"]), reverse=True))

    #sorted_scores_list = [float(values["impaakt_score"]) for values in sorted_sources.values()]

    # Afficher le résultat trié et la liste des valeurs triées
    #for url, values in sorted_sources.items():
    #    print(f"URL: {url}, Impaakt Score: {values['impaakt_score']}")



    # New sheet, set as active sheet
    sheet = workbook.create_sheet(i_str)
    workbook.active = workbook[i_str]
    sheet = workbook.active

    # Set column dimensions
    for col, width in zip('ABCDEF', [50, 10, 10, 50, 10, 10]):
        sheet.column_dimensions[col].width = width

    # Set headers and styles
    sheet["A1"] = query
    sheet["A3"] = "Ranking A"
    sheet["B3"] = "Accessible"
    sheet["C3"] = "Relevant"
    sheet["D3"] = "Ranking B"
    sheet["E3"] = "Accessible"
    sheet["F3"] = "Relevant"
    sheet['A1'].fill = PatternFill("solid", start_color="FFC000")
    sheet["A3"].fill,sheet["B3"].fill,sheet["C3"].fill = pattern_green,pattern_green,pattern_green
    sheet["D3"].fill,sheet["E3"].fill,sheet["F3"].fill = pattern_orange,pattern_orange,pattern_orange
    sheet["A3"].font,sheet["B3"].font,sheet["C3"].font,sheet["D3"].font,sheet["E3"].font, sheet["F3"].font = font_white,font_white,font_white,font_white,font_white,font_white

    j=4
    
    # Apply the validation rules to specific cells
    col = 'B'
    dv.add(f'{col}{j}:{col}{len(sources)}')
    col = 'C'
    dv.add(f'{col}{j}:{col}{len(sources)}')
    col = 'E'
    dv.add(f'{col}{j}:{col}{len(sources)}')
    col = 'F'
    dv.add(f'{col}{j}:{col}{len(sources)}')
    # Add the validation rules to the sheet
    sheet.add_data_validation(dv)
    
    for url,url_2 in zip(sources, sorted_sources) :
        sheet.cell(row=j, column=1).hyperlink = url
        sheet.cell(row=j, column=1).style = "Hyperlink"
        sheet.cell(row=j, column=1).alignment = Alignment(wrap_text=True)
        sheet.cell(row=j, column=4).hyperlink = url_2
        sheet.cell(row=j, column=4).style = "Hyperlink"
        sheet.cell(row=j, column=4).alignment = Alignment(wrap_text=True)
        j=j+1

    print(i)
    i += 1

# save the data
workbook.save(filename=filename)
